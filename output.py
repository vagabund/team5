from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Side, Border
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from constants import PLAYED_VARS, output_path

CENTER = Alignment(horizontal="center", vertical="center")
RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF")
PLAYED_FILL = PatternFill(fill_type="solid", fgColor="FFFF00")
PLAYED_FONT = Font(bold=True)
COMMON_FILL = PatternFill(fill_type="solid", fgColor="D9D9D9")
THIN = Side(style="thin", color="000000")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def create_xlsx(results: list, df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Результаты"
    next_row = draw_markers(ws, 1, 4)
    current_row = next_row + 1
    for i in results:
        current_row = draw_match(ws, current_row, i)
    ws_squad = wb.create_sheet("Сборная")
    for r in dataframe_to_rows(df, index=False, header=True):
        ws_squad.append(r)
    for cell in ws_squad[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws_squad.column_dimensions["A"].width = 25
    ws_squad.column_dimensions["B"].width = 20
    wb.save(output_path / "output.xlsx")

def draw_markers(ws, start_row: int = 1, start_col: int = 4):
    markers_per_row = 10

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["R"].width = 20

    marker = 1
    row = start_row
    for _ in range(4):
        for i in range(markers_per_row):
            col = start_col + i
            cell = ws.cell(row, col, marker)
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 3
            cell.alignment = CENTER
            if marker in PLAYED_VARS:
                cell.fill = PLAYED_FILL
                cell.font = PLAYED_FONT
            marker += 1
        row += 1

    return row


def draw_player_row(ws, row: int, pair: dict):
    home_player = pair["home_player"]
    away_player = pair["away_player"]
    initial_home = sorted(pair["initial_home"])
    initial_away = sorted(pair["initial_away"])
    home_vars = set(pair["home_vars"])
    away_vars = set(pair["away_vars"])
    home_score, away_score = pair["score"]
    common_vars = set(initial_home) & set(initial_away)

    cell_home_player = ws.cell(row=row, column=1, value=home_player)
    cell_away_player = ws.cell(row=row, column=18, value=away_player)
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["R"].width = 25
    cell_home_player.alignment = RIGHT_ALIGN
    cell_away_player.alignment = LEFT_ALIGN


    for i, val in enumerate(initial_home[:7]):
        col = 2 + i
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 3
        cell = ws.cell(row=row, column=col, value=val)
        cell.alignment = CENTER
        # cell.border = THIN_BORDER
        if val in home_vars:
            cell.fill = PLAYED_FILL
            cell.font = PLAYED_FONT
        elif val in common_vars:
            cell.fill = COMMON_FILL
    col_letter = get_column_letter(9)
    ws.column_dimensions[col_letter].width = 3
    col_letter = get_column_letter(10)
    ws.column_dimensions[col_letter].width = 3
    cell_home_score = ws.cell(row=row, column=9, value=home_score)
    cell_away_score = ws.cell(row=row, column=10, value=away_score)
    cell_home_score.alignment = CENTER
    cell_away_score.alignment = CENTER
    # cell_home_score.border = THIN_BORDER
    # cell_away_score.border = THIN_BORDER
    cell_home_score.font = PLAYED_FONT
    cell_away_score.font = PLAYED_FONT

    for i, val in enumerate(initial_away[:7]):
        col = 11 + i
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 3
        cell = ws.cell(row=row, column=col, value=val)
        cell.alignment = CENTER
        # cell.border = THIN_BORDER
        if val in away_vars:
            cell.fill = PLAYED_FILL
            cell.font = PLAYED_FONT
        elif val in common_vars:
            cell.fill = COMMON_FILL

def draw_match(ws, start_row: int, match: dict):
    home_name = match["home"]
    away_name = match["away"]
    home_goals = match["home_goals"]
    away_goals = match["away_goals"]
    home_captain = match["home_captain"]
    away_captain = match["away_captain"]
    pairs = match["pairs"]

    home_total_guessed = 0
    away_total_guessed = 0
    for i in pairs:
        home_total_guessed += len(i['home_overall'])
        away_total_guessed += len(i['away_overall'])

    header_row = start_row

    cell_home_team = ws.cell(row=header_row, column=1, value=home_name)
    cell_away_team = ws.cell(row=header_row, column=18, value=away_name)
    cell_home_team.alignment = RIGHT_ALIGN
    cell_away_team.alignment = LEFT_ALIGN
    cell_home_team.font = HEADER_FONT
    cell_away_team.font = HEADER_FONT
    cell_home_team.fill = HEADER_FILL
    cell_away_team.fill = HEADER_FILL
    cell_home_goals = ws.cell(row=header_row, column=9, value=home_goals)
    cell_away_goals = ws.cell(row=header_row, column=10, value=away_goals)
    cell_home_goals.alignment = CENTER
    cell_away_goals.alignment = CENTER
    cell_home_goals.fill = HEADER_FILL
    cell_away_goals.fill = HEADER_FILL
    cell_home_goals.font = HEADER_FONT
    cell_away_goals.font = HEADER_FONT

    trainer_row = header_row + 1
    cell_home_captain = ws.cell(row=trainer_row, column=1, value=home_captain)
    cell_away_captain = ws.cell(row=trainer_row, column=18, value=away_captain)
    cell_home_captain.font = HEADER_FONT
    cell_away_captain.font = HEADER_FONT
    cell_home_captain.fill = HEADER_FILL
    cell_away_captain.fill = HEADER_FILL
    cell_home_captain.alignment = RIGHT_ALIGN
    cell_away_captain.alignment = LEFT_ALIGN
    cell_home_total = ws.cell(row=trainer_row, column=9, value=home_total_guessed)
    cell_away_total = ws.cell(row=trainer_row, column=10, value=away_total_guessed)
    cell_home_total.alignment = CENTER
    cell_away_total.alignment = CENTER
    cell_home_total.font = PLAYED_FONT
    cell_away_total.font = PLAYED_FONT

    row = trainer_row + 1
    for pair in pairs:
        draw_player_row(ws, row=row, pair=pair)
        row += 1

    next_start_row = row + 1
    return next_start_row